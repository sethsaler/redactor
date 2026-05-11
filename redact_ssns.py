#!/usr/bin/env python3
"""Redact sensitive information from documents.

Supports PDF, CSV, XLSX, and TXT files.
Modes:
  - ssn: Redact Social Security Numbers (default, backward compatible)
  - bank: Redact numbers (except currency amounts), emails, phones, names, addresses
  - all: Redact both SSNs and bank patterns

For bank statements, currency amounts (e.g., $1,234.56, €500) are preserved.

Optional custom phrases (CLI --phrase / --phrases-file, or GUI) use a separate
pass: when any phrase is given, only those literals are redacted (SSN/bank/all
patterns are skipped). Comma-separated lists are split into separate phrases.
Matching is literal and case-insensitive.
"""

import argparse
import csv
import io
import re
import sys
from itertools import groupby
from pathlib import Path

# PDF support
try:
    import fitz  # PyMuPDF

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Image/OCR support
try:
    from PIL import Image

    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

OCR_AVAILABLE = False
try:
    import pytesseract
    from pdf2image import convert_from_path

    OCR_AVAILABLE = True
except ImportError:
    pass

# Excel support
try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ============================================================================
# PATTERN DEFINITIONS
# ============================================================================

# SSN patterns: 123-45-6789 and 123456789
SSN_PATTERNS = [
    re.compile(r"\b(?!000|666|9\d{2})\d{3}-(?!00)\d{2}-(?!0000)\d{4}\b"),
    re.compile(r"\b(?!000|666|9\d{2})\d{3}(?!00)\d{2}(?!0000)\d{4}\b"),
]

SSN_INVALID_AREA = set(str(i) for i in range(734, 750))

# Currency symbols to preserve
CURRENCY_SYMBOLS = r"$€£¥₹₩¢"
CURRENCY_CHARS = list(CURRENCY_SYMBOLS)

# Number pattern: matches 4+ digit sequences
# Requires at least 4 consecutive digits (ignoring common separators)
# We'll filter out matches preceded by currency in the find function
BANK_NUMBER_PATTERN = re.compile(
    r"(?<![\d\-])"  # Not preceded by another digit or hyphen
    r"\b\d{4,}\b"  # 4+ consecutive digits
    r"|\b\d{1,3}(?:,\d{3})+\b"  # OR comma-separated thousands like 1,234,567
    r"|\b\d{1,3}(?:\.\d{3})+\b"  # OR dot-separated thousands (European)
)

# Email pattern
EMAIL_PATTERN = re.compile(
    r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b", re.IGNORECASE
)

# Phone patterns (various international formats)
PHONE_PATTERNS = [
    re.compile(r"\b\+?1?[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}\b"),  # US/Canada
    re.compile(
        r"\b\+?\d{1,3}[-.\s]?\d{1,4}[-.\s]?\d{1,4}[-.\s]?\d{1,9}\b"
    ),  # International
    re.compile(r"\b\d{3}[-.\s]?\d{3}[-.\s]?\d{4}\b"),  # Simple 10-digit
]

# Name patterns (heuristic approach)
# Looks for capitalized words that appear to be names
NAME_PATTERNS = [
    # Full name: "John Smith" or "J. Smith" or "Smith, John"
    re.compile(
        r"\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)+\b"
    ),  # "John Smith", "Mary Jane Watson"
    re.compile(r"\b[A-Z]\.\s*[A-Z][a-z]+\b"),  # "J. Smith"
    re.compile(r"\b[A-Z][a-z]+,\s*[A-Z][a-z]+\b"),  # "Smith, John"
    re.compile(r"\b[A-Z][a-z]+,\s*[A-Z]\.\s*[A-Z][a-z]+\b"),  # "Smith, J. R."
]

# Common name indicators to increase confidence
NAME_INDICATORS = ["mr", "mrs", "ms", "miss", "dr", "prof", "sir", "madam"]

# Address patterns
ADDRESS_PATTERNS = [
    # Street addresses: "123 Main St", "456 Oak Avenue, Apt 2B"
    re.compile(
        r"\b\d+\s+[A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)*(?:\s+(?:St|Street|Ave|Avenue|Rd|Road|Dr|Drive|Blvd|Boulevard|Ln|Lane|Way|Ct|Court|Pl|Place|Circle|Cir|Trail|Trl|Parkway|Pkwy|Hwy|Highway|Route|Rte))\b",
        re.IGNORECASE,
    ),
    # City, State ZIP
    re.compile(r"\b[A-Z][a-z]+(?:\s[A-Z][a-z]+)?,\s*[A-Za-z\s]{2,}\s*\d{5}(-\d{4})?\b"),
    # ZIP code alone (5 or 9 digit)
    re.compile(r"\b\d{5}(?:-\d{4})?\b"),
]

# Aggregation of all bank patterns
BANK_PATTERNS = [
    ("number", BANK_NUMBER_PATTERN),
    ("email", EMAIL_PATTERN),
    ("phone", PHONE_PATTERNS),
    ("name", NAME_PATTERNS),
    ("address", ADDRESS_PATTERNS),
]

MIN_TEXT_LENGTH = 20


# ============================================================================
# CUSTOM PHRASES (user-supplied literals)
# ============================================================================


def normalize_phrase_list(phrases):
    """Dedupe phrases while preserving order; strips whitespace. Case-insensitive dedup."""
    if not phrases:
        return []
    seen = set()
    out = []
    for p in phrases:
        if not p or not isinstance(p, str):
            continue
        s = p.strip()
        if not s:
            continue
        key = s.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


def parse_user_phrase_inputs(raw_strings):
    """Split comma-separated entries and normalize (order preserved, deduped)."""
    expanded = []
    for raw in raw_strings or []:
        if not isinstance(raw, str):
            continue
        raw = raw.strip()
        if not raw:
            continue
        for part in raw.split(","):
            p = part.strip()
            if p:
                expanded.append(p)
    return normalize_phrase_list(expanded)


def load_phrases_from_file(path):
    """Load newline-separated phrases; skip blanks and # comments."""
    path = Path(path)
    text = path.read_text(encoding="utf-8", errors="ignore")
    out = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        out.append(line)
    return out


def find_custom_phrases_in_text(text, phrases):
    """Find literal occurrences of user-supplied phrases (case-insensitive)."""
    if not phrases:
        return []
    matches = []
    seen_spans = set()
    for phrase in phrases:
        if not phrase or not isinstance(phrase, str):
            continue
        stripped = phrase.strip()
        if not stripped:
            continue
        pat = re.compile(re.escape(stripped), re.IGNORECASE)
        for m in pat.finditer(text):
            span = (m.start(), m.end())
            if span in seen_spans:
                continue
            seen_spans.add(span)
            matches.append((m.group(), m.start(), m.end(), "custom"))
    return matches


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================


def _normalize(text):
    return text.replace("-", "").replace(" ", "")


def is_valid_ssn(text):
    """Validate SSN against SSA rules."""
    normalized = _normalize(text)
    if len(normalized) != 9 or not normalized.isdigit():
        return False
    if normalized in ("000000000", "666000000"):
        return False
    area = normalized[:3]
    if area.startswith("0") or area.startswith("666") or area.startswith("9"):
        return False
    if area in SSN_INVALID_AREA:
        return False
    group = normalized[3:5]
    serial = normalized[5:]
    if group == "00" or serial == "0000":
        return False
    return True


# ============================================================================
# TEXT DETECTION FUNCTIONS
# ============================================================================


def find_ssns_in_text(text):
    """Find all valid SSNs in text."""
    ssns = []
    seen = set()
    for pattern in SSN_PATTERNS:
        for m in pattern.finditer(text):
            raw = m.group()
            normalized = _normalize(raw)
            if normalized in seen:
                continue
            if is_valid_ssn(raw):
                seen.add(normalized)
                ssns.append((raw, m.start(), m.end()))
    return ssns


def _is_preceded_by_currency(text, start_pos):
    """Check if the position is preceded by a currency symbol (with optional space)."""
    if start_pos == 0:
        return False
    # Check for currency symbol directly before
    if start_pos > 0 and text[start_pos - 1] in CURRENCY_CHARS:
        return True
    # Check for currency symbol with one space before
    if (
        start_pos > 1
        and text[start_pos - 2] in CURRENCY_CHARS
        and text[start_pos - 1].isspace()
    ):
        return True
    return False


def _looks_like_ssn(text):
    """Check if text looks like an SSN (9 digits with optional separators)."""
    normalized = (
        text.replace("-", "").replace(" ", "").replace(",", "").replace(".", "")
    )
    return len(normalized) == 9 and normalized.isdigit()


def _looks_like_decimal_amount(text):
    """Check if text looks like a decimal amount (e.g., 1,234.56 or 1.234,56)."""
    # Pattern: digits followed by comma/dot then exactly 2 digits (cents)
    if re.search(r"[\d][\.,]\d{2}$", text):
        return True
    return False


def find_bank_patterns_in_text(text):
    """Find all bank-sensitive patterns in text (numbers, emails, phones, names, addresses)."""
    matches = []
    seen_spans = set()  # Track to avoid overlapping matches

    # Numbers (excluding currency amounts, SSNs, and decimal amounts)
    for m in BANK_NUMBER_PATTERN.finditer(text):
        span = (m.start(), m.end())
        if span in seen_spans:
            continue
        raw = m.group()
        # Skip if preceded by currency symbol
        if _is_preceded_by_currency(text, m.start()):
            continue
        # Skip if looks like an SSN (let SSN patterns handle those)
        if _looks_like_ssn(raw):
            continue
        # Skip if looks like a decimal amount (e.g., 1,234.56)
        if _looks_like_decimal_amount(raw):
            continue
        seen_spans.add(span)
        matches.append((raw, m.start(), m.end(), "number"))

    # Emails
    for m in EMAIL_PATTERN.finditer(text):
        span = (m.start(), m.end())
        if span not in seen_spans:
            seen_spans.add(span)
            matches.append((m.group(), m.start(), m.end(), "email"))

    # Phones (exclude SSN-like patterns)
    for patterns in PHONE_PATTERNS:
        for m in patterns.finditer(text):
            span = (m.start(), m.end())
            if span in seen_spans:
                continue
            raw = m.group()
            # Skip if looks like an SSN (to avoid matching SSNs as phone numbers)
            if _looks_like_ssn(raw):
                continue
            seen_spans.add(span)
            matches.append((raw, m.start(), m.end(), "phone"))

    # Names
    for pattern in NAME_PATTERNS:
        for m in pattern.finditer(text):
            span = (m.start(), m.end())
            if span not in seen_spans:
                seen_spans.add(span)
                matches.append((m.group(), m.start(), m.end(), "name"))

    # Addresses
    for pattern in ADDRESS_PATTERNS:
        for m in pattern.finditer(text):
            span = (m.start(), m.end())
            if span not in seen_spans:
                seen_spans.add(span)
                matches.append((m.group(), m.start(), m.end(), "address"))

    return matches


def find_all_patterns_in_text(text):
    """Find both SSNs and bank patterns."""
    ssns = find_ssns_in_text(text)
    bank = find_bank_patterns_in_text(text)

    # Convert SSNs to same format
    ssn_matches = [(raw, start, end, "ssn") for raw, start, end in ssns]

    return ssn_matches + bank


def find_matches_in_text(text, mode="ssn", custom_phrases=None):
    """Find matches: if custom_phrases is non-empty, only those literals; else mode patterns."""
    custom_phrases = normalize_phrase_list(custom_phrases or [])
    if custom_phrases:
        return find_custom_phrases_in_text(text, custom_phrases)

    if mode == "ssn":
        return [(raw, start, end, "ssn") for raw, start, end in find_ssns_in_text(text)]
    elif mode == "bank":
        return find_bank_patterns_in_text(text)
    elif mode == "all":
        return find_all_patterns_in_text(text)
    return []


# ============================================================================
# PDF REDACTION
# ============================================================================


def redact_text_page(page, mode="ssn", custom_phrases=None):
    """Redact matches from a text-based PDF page."""
    count = 0
    text = page.get_text("text")
    matches = find_matches_in_text(text, mode, custom_phrases)

    for raw, start, end, pattern_type in matches:
        instances = page.search_for(raw)
        for rect in instances:
            page.add_redact_annot(rect, fill=(0, 0, 0))
            count += 1

    if count > 0:
        page.apply_redactions()

    return count


def _redact_ocr_lines_custom_phrases(
    page, img_w, img_h, page_rect, data, custom_phrases
):
    """Redact custom phrases on OCR output by matching within each text line."""
    scale_x = page_rect.width / img_w
    scale_y = page_rect.height / img_h
    count = 0
    n_boxes = len(data["text"])

    row_indices = []
    for i in range(n_boxes):
        if not (data["text"][i] or "").strip():
            continue
        row_indices.append(i)

    row_indices.sort(
        key=lambda i: (
            data["block_num"][i],
            data["par_num"][i],
            data["line_num"][i],
            data["word_num"][i],
        )
    )

    for _key, grp in groupby(
        row_indices,
        key=lambda i: (
            data["block_num"][i],
            data["par_num"][i],
            data["line_num"][i],
        ),
    ):
        idxs = list(grp)
        words = [(data["text"][i] or "").strip() for i in idxs]
        line_text = " ".join(words)
        if not line_text:
            continue

        word_ranges = []
        pos = 0
        for wi, w in enumerate(words):
            s = pos
            e = pos + len(w)
            word_ranges.append((s, e))
            pos = e
            if wi < len(words) - 1:
                pos += 1

        for phrase in custom_phrases:
            pat = re.compile(re.escape(phrase), re.IGNORECASE)
            for m in pat.finditer(line_text):
                ms, me = m.span()
                min_x = min_y = None
                max_x = max_y = None
                for j, (ws, we) in enumerate(word_ranges):
                    if we <= ms or ws >= me:
                        continue
                    i = idxs[j]
                    x = data["left"][i]
                    y = data["top"][i]
                    w = data["width"][i]
                    h = data["height"][i]
                    if min_x is None:
                        min_x, min_y, max_x, max_y = x, y, x + w, y + h
                    else:
                        min_x = min(min_x, x)
                        min_y = min(min_y, y)
                        max_x = max(max_x, x + w)
                        max_y = max(max_y, y + h)
                if min_x is None:
                    continue

                pdf_x0 = min_x * scale_x
                pdf_y0 = min_y * scale_y
                pdf_x1 = max_x * scale_x
                pdf_y1 = max_y * scale_y
                pad_w = (pdf_x1 - pdf_x0) * 0.15
                pad_h = (pdf_y1 - pdf_y0) * 0.15
                pad = max(pad_w, pad_h, 3)
                rect = fitz.Rect(
                    pdf_x0 - pad,
                    pdf_y0 - pad,
                    pdf_x1 + pad,
                    pdf_y1 + pad,
                )
                page.add_redact_annot(rect, fill=(0, 0, 0))
                count += 1

    return count


def redact_image_page(page, page_num, mode="ssn", dpi=300, custom_phrases=None):
    """Redact matches from an image-based PDF page using OCR."""
    if not OCR_AVAILABLE:
        return 0

    custom_norm = normalize_phrase_list(custom_phrases or [])

    count = 0
    try:
        # Get the document from the page and create a temporary file
        doc = page.parent
        with io.BytesIO() as pdf_bytes:
            doc.save(pdf_bytes)
            pdf_bytes.seek(0)

            images = convert_from_path(
                pdf_bytes,
                first_page=page_num + 1,
                last_page=page_num + 1,
                dpi=dpi,
            )
    except Exception:
        return 0

    if not images:
        return 0

    img = images[0]
    img_w, img_h = img.size
    page_rect = page.rect
    scale_x = page_rect.width / img_w
    scale_y = page_rect.height / img_h

    try:
        data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
    except pytesseract.TesseractNotFoundError:
        print("  Warning: Tesseract OCR not installed. Skipping image-based page.")
        return 0

    if custom_norm:
        count = _redact_ocr_lines_custom_phrases(
            page, img_w, img_h, page_rect, data, custom_norm
        )
    else:
        n_boxes = len(data["text"])
        for i in range(n_boxes):
            word = data["text"][i].strip()
            if not word:
                continue

            matches = find_matches_in_text(word, mode, None)
            if not matches:
                continue

            x = data["left"][i]
            y = data["top"][i]
            w = data["width"][i]
            h = data["height"][i]

            pdf_x = x * scale_x
            pdf_y = y * scale_y
            pdf_w = w * scale_x
            pdf_h = h * scale_y

            padding = max(pdf_w * 0.15, 3)
            rect = fitz.Rect(
                pdf_x - padding,
                pdf_y - padding,
                pdf_x + pdf_w + padding,
                pdf_y + pdf_h + padding,
            )

            page.add_redact_annot(rect, fill=(0, 0, 0))
            count += 1

    if count > 0:
        page.apply_redactions()

    return count


def is_text_page(page):
    """Determine if a page is text-based or image-based."""
    text = page.get_text("text").strip()
    return len(text) >= MIN_TEXT_LENGTH


def redact_pdf(pdf_path, output_path=None, mode="ssn", verbose=False, custom_phrases=None):
    """Redact a PDF file."""
    if not PDF_AVAILABLE:
        print("Error: PyMuPDF not installed. Cannot process PDF.")
        return 0

    pdf_path = Path(pdf_path)
    if output_path is None:
        output_path = pdf_path.with_name(f"{pdf_path.stem}_redacted{pdf_path.suffix}")

    doc = fitz.open(str(pdf_path))
    total_redacted = 0

    for page_num in range(len(doc)):
        page = doc[page_num]

        if is_text_page(page):
            count = redact_text_page(page, mode=mode, custom_phrases=custom_phrases)
            if verbose and count:
                print(f"  Page {page_num + 1}: redacted {count} item(s) (text)")
            total_redacted += count
        else:
            count = redact_image_page(
                page, page_num, mode=mode, custom_phrases=custom_phrases
            )
            if verbose and count:
                print(f"  Page {page_num + 1}: redacted {count} item(s) (OCR)")
            total_redacted += count

    doc.save(str(output_path))
    doc.close()

    return total_redacted


# ============================================================================
# TEXT FILE REDACTION
# ============================================================================


def redact_text_file(text_path, output_path=None, mode="ssn", custom_phrases=None):
    """Redact a plain text file using regex substitution."""
    text_path = Path(text_path)
    if output_path is None:
        output_path = text_path.with_name(
            f"{text_path.stem}_redacted{text_path.suffix}"
        )

    content = text_path.read_text(encoding="utf-8", errors="ignore")
    redaction_count = 0

    matches = find_matches_in_text(content, mode, custom_phrases)
    for raw, start, end, _ in sorted(matches, key=lambda x: x[1], reverse=True):
        content = content[:start] + "[REDACTED]" + content[end:]
        redaction_count += 1

    output_path.write_text(content, encoding="utf-8")
    return redaction_count


# ============================================================================
# CSV REDACTION
# ============================================================================


def redact_csv_file(csv_path, output_path=None, mode="ssn", custom_phrases=None):
    """Redact a CSV file by processing cell values."""
    csv_path = Path(csv_path)
    if output_path is None:
        output_path = csv_path.with_name(f"{csv_path.stem}_redacted{csv_path.suffix}")

    redaction_count = 0

    with open(csv_path, "r", encoding="utf-8", newline="", errors="ignore") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        # Empty file
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            pass
        return 0

    redacted_rows = []
    for row in rows:
        redacted_row = []
        for cell in row:
            redacted_cell, count = redact_cell(cell, mode, custom_phrases)
            redaction_count += count
            redacted_row.append(redacted_cell)
        redacted_rows.append(redacted_row)

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(redacted_rows)

    return redaction_count


def redact_cell(cell, mode="ssn", custom_phrases=None):
    """Redact sensitive patterns in a single cell value."""
    if not cell or not isinstance(cell, str):
        return cell, 0

    matches = find_matches_in_text(cell, mode, custom_phrases)
    result = cell
    redaction_count = 0
    for raw, start, end, _ in sorted(matches, key=lambda x: x[1], reverse=True):
        result = result[:start] + "[REDACTED]" + result[end:]
        redaction_count += 1
    return result, redaction_count


# ============================================================================
# EXCEL REDACTION
# ============================================================================


def redact_excel_file(excel_path, output_path=None, mode="ssn", custom_phrases=None):
    """Redact an Excel (.xlsx) file by processing cell values."""
    if not EXCEL_AVAILABLE:
        print("Error: openpyxl not installed. Cannot process Excel files.")
        return 0

    excel_path = Path(excel_path)
    if output_path is None:
        output_path = excel_path.with_name(
            f"{excel_path.stem}_redacted{excel_path.suffix}"
        )

    wb = openpyxl.load_workbook(str(excel_path))
    redaction_count = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    redacted_value, count = redact_cell(
                        cell.value, mode, custom_phrases
                    )
                    redaction_count += count
                    cell.value = redacted_value

    wb.save(str(output_path))
    return redaction_count


# ============================================================================
# FILE TYPE DISPATCH
# ============================================================================


def detect_file_type(file_path):
    """Detect file type based on extension."""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return "pdf"
    elif ext == ".csv":
        return "csv"
    elif ext == ".txt":
        return "txt"
    elif ext in (".xlsx", ".xlsm"):
        return "excel"
    elif ext == ".xls":
        return "unsupported"  # Legacy Excel not supported
    return "unknown"


def redact_file(file_path, output_path=None, mode="ssn", verbose=False, custom_phrases=None):
    """Redact a single file based on its type."""
    file_path = Path(file_path)
    file_type = detect_file_type(file_path)

    if file_type == "pdf":
        return redact_pdf(file_path, output_path, mode, verbose, custom_phrases)
    elif file_type == "txt":
        return redact_text_file(file_path, output_path, mode, custom_phrases)
    elif file_type == "csv":
        return redact_csv_file(file_path, output_path, mode, custom_phrases)
    elif file_type == "excel":
        return redact_excel_file(file_path, output_path, mode, custom_phrases)
    elif file_type == "unsupported":
        print(
            f"  Skipping {file_path.name}: Legacy .xls format not supported (convert to .xlsx)"
        )
        return 0
    else:
        print(f"  Skipping {file_path.name}: Unknown file type")
        return 0


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================


def cli_main():
    """Command-line interface for redaction."""
    parser = argparse.ArgumentParser(
        description="Redact sensitive information from documents (PDF, CSV, TXT, XLSX).",
        epilog="""
Modes (used only when no phrases are given):
  ssn  - Redact Social Security Numbers only (default)
  bank - Redact numbers (except currency amounts), emails, phones, names, addresses
  all  - Redact both SSNs and bank patterns

When --phrase or --phrases-file is set, only those literals are redacted (mode is ignored).
Comma-separated values split into separate phrases. Currency amounts (e.g. $1,234.56, €500)
are preserved in bank mode.

Examples:
  python redact_ssns.py document.pdf
  python redact_ssns.py statement.pdf --mode bank
  python redact_ssns.py data/ --recursive --mode all -o ./redacted/
  python redact_ssns.py memo.txt --phrase "Project X, Acme Corp"
  python redact_ssns.py report.pdf --phrases-file ./phrases.txt
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "path",
        type=Path,
        nargs="?",  # Make path optional when launching GUI
        help="File or directory to process (PDF, CSV, TXT, XLSX)",
    )
    parser.add_argument(
        "-r", "--recursive", action="store_true", help="Recurse into subdirectories"
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="Log each item found"
    )
    parser.add_argument(
        "-o", "--output-dir", type=Path, default=None, help="Output directory"
    )
    parser.add_argument(
        "-m",
        "--mode",
        choices=["ssn", "bank", "all"],
        default="ssn",
        help="Redaction mode when no phrases are given (default: ssn)",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Launch GUI mode (default when no arguments provided)",
    )
    parser.add_argument(
        "-p",
        "--phrase",
        action="append",
        default=[],
        metavar="TEXT",
        help="Word or phrase to redact only (repeatable); comma-separated splits into multiple; case-insensitive literal match; if set, mode patterns are skipped",
    )
    parser.add_argument(
        "--phrases-file",
        type=Path,
        default=None,
        metavar="PATH",
        help="Phrases file: each line is one or more comma-separated phrases (# starts a comment); if set, mode patterns are skipped",
    )

    args = parser.parse_args()

    # Check if we should launch GUI
    if args.gui or args.path is None:
        launch_gui()
        return

    target = args.path
    mode = args.mode

    chunks = list(args.phrase)
    if args.phrases_file:
        if not args.phrases_file.is_file():
            print(f"Phrases file not found: {args.phrases_file}")
            sys.exit(1)
        chunks.extend(load_phrases_from_file(args.phrases_file))
    custom_phrases = parse_user_phrase_inputs(chunks)

    # Check dependencies based on file types we'll process
    if not PDF_AVAILABLE:
        print("Warning: PyMuPDF not installed. PDF files will be skipped.")
        print("Install with: pip install PyMuPDF")
    if not EXCEL_AVAILABLE:
        print("Warning: openpyxl not installed. Excel files will be skipped.")
        print("Install with: pip install openpyxl")
    if not OCR_AVAILABLE:
        print("Warning: OCR dependencies missing. Scanned PDFs will be skipped.")
        print("Install with: pip install pytesseract pdf2image Pillow")
        print("Also install Tesseract: https://github.com/tesseract-ocr/tesseract")
    if not any([PDF_AVAILABLE, EXCEL_AVAILABLE]):
        print()

    # Collect files to process
    supported_patterns = ["*.pdf", "*.csv", "*.txt", "*.xlsx", "*.xlsm"]
    files = []

    if target.is_file():
        files = [target]
    elif target.is_dir():
        if args.recursive:
            for pattern in supported_patterns:
                files.extend(target.rglob(pattern))
        else:
            for pattern in supported_patterns:
                files.extend(target.glob(pattern))
        files = sorted(set(files))
        files = [f for f in files if "_redacted" not in f.stem]
        if not files:
            print(f"No supported files found in {target}")
            sys.exit(1)
    else:
        print(f"Path not found: {target}")
        sys.exit(1)

    # Process files
    total_files = 0
    total_redactions = 0

    for file in files:
        output = None
        if args.output_dir:
            args.output_dir.mkdir(parents=True, exist_ok=True)
            output = args.output_dir / f"{file.stem}_redacted{file.suffix}"

        if args.verbose:
            print(f"Processing: {file}")

        count = redact_file(
            file,
            output_path=output,
            mode=mode,
            verbose=args.verbose,
            custom_phrases=custom_phrases,
        )
        total_files += 1
        total_redactions += count

        # Determine final output path
        if output:
            final_output = output
        else:
            final_output = file.with_name(f"{file.stem}_redacted{file.suffix}")

        status = f"Redacted {count} item(s)" if count else "No sensitive data found"
        print(f"  {file.name}: {status} -> {final_output}")

    mode_desc = {"ssn": "SSN(s)", "bank": "bank-sensitive item(s)", "all": "item(s)"}
    if custom_phrases:
        desc = "listed phrase occurrence(s)"
    else:
        desc = mode_desc[mode]
    print(
        f"\nDone: {total_files} file(s) processed, {total_redactions} {desc} redacted."
    )


def launch_gui():
    """Launch the GUI interface."""
    try:
        from tkinter import Tk
        from redactor_gui import RedactorGUI

        root = Tk()
        app = RedactorGUI(root)
        app.run()
    except ImportError as e:
        print(f"Error: GUI dependencies not available. {e}")
        print("Make sure redactor_gui.py is in the same directory.")
        sys.exit(1)
    except Exception as e:
        print(f"Error launching GUI: {e}")
        sys.exit(1)


def main():
    """Main entry point - launches GUI if no args, otherwise CLI."""
    # Check if any arguments provided (excluding script name)
    if len(sys.argv) == 1:
        # No arguments - launch GUI
        launch_gui()
    else:
        # Arguments provided - use CLI
        cli_main()


if __name__ == "__main__":
    main()
